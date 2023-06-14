import grequests
import pandas as pd
import threading
from time import sleep
import datetime
import openpyxl
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox as ms
from pathlib import Path
import os
import json
import sys

#initializing variables
pd.options.display.max_rows = 1000
refreshRate=1

autoIncrement=0
userName=None
startProcess=False
stopProcess=False
runningGui=None
exitDay=datetime.datetime.now()

#To Retrieve expiry date
# def getExpiryDate():
#     global exitDay
#     try:
#         expPath = Path(r'C:\Test')
#         if not os.path.exists(expPath):
#             os.makedirs(expPath)
#             os.system(f'attrib +h "{expPath}"')
#         file_path = expPath.joinpath('sysed.txt')
#         if file_path.is_file():
#             with open (file_path,'r') as f:
#                 expiryString= str(f.read()).split("/")
#                 exitDay=exitDay.replace(day=int(expiryString[0]),month=int(expiryString[1]),year=int(expiryString[2]),hour=23,minute=45,second=0)
#         else:
#             exitDay = exitDay + datetime.timedelta(days=6)
#             exitDay=exitDay.replace(hour=23,minute=45,second=0)
#             with open (file_path,'w') as f:
#                 f.write(str(exitDay.day)+"/"+str(exitDay.month)+"/"+str(exitDay.year))
#     except:
#         exitDay=exitDay.replace(day=15,month=3,year=2045,hour=23,minute=45,second=0)

# getExpiryDate()

# if(datetime.datetime.now()>=exitDay):
#     #print("Trading subscription got expired")
#     sys.exit()
   

#URL to store reports and donelist in some path
newpath=r'C:\Temp\stooxoreport'
dirpath = Path(newpath)
if not os.path.exists(newpath):
    os.makedirs(newpath)

#URL to read csv files
excelurl=r'C:\Users\Dell\Desktop\Auto Trade'
   
#initializing global variables
async_list = []

equityGroups={}
futureGroups={}
optionGroups={}
indexFutureGroups={}
indexOptionGroups={}
commodityGroups={}
NSETiming={"start":{"hour":9,"min":16},"end":{"hour":15,"min":25}}
MCXTiming={"start":{"hour":9,"min":1},"end":{"hour":23,"min":25}}

equityDone={"BUY":{},"SELL":{},"EXIT BUY":{},"EXIT SELL":{}}
futureDone={"BUY":{},"SELL":{},"EXIT BUY":{},"EXIT SELL":{}}
optionDone={"BUY":{},"SELL":{},"EXIT BUY":{},"EXIT SELL":{}}
indexFutureDone={"BUY":{},"SELL":{},"EXIT BUY":{},"EXIT SELL":{}}
indexOptionDone={"BUY":{},"SELL":{},"EXIT BUY":{},"EXIT SELL":{}}
commodityDone={"BUY":{},"SELL":{},"EXIT BUY":{},"EXIT SELL":{}}

equityDoneRow=0
futureDoneRow=0
optionDoneRow=0
indexFutureDoneRow=0
indexOptionDoneRow=0
commodityDoneRow=0

equityUpdated=datetime.datetime.timestamp(datetime.datetime.now())
futureUpdated=datetime.datetime.timestamp(datetime.datetime.now())
optionUpdated=datetime.datetime.timestamp(datetime.datetime.now())
indexFutureUpdated=datetime.datetime.timestamp(datetime.datetime.now())
indexOptionUpdated=datetime.datetime.timestamp(datetime.datetime.now())
commodityUpdated=datetime.datetime.timestamp(datetime.datetime.now())


#To initialize inceament
def initialiseIncrementValue():
    global autoIncrement
    global dirpath
    file_path = dirpath.joinpath('increment.txt')
    if not file_path.is_file():
        with open (file_path,'w') as f:  
            f.write(str(autoIncrement))
            #print('File was created for auto increment')
initialiseIncrementValue()

#To store the unique id in local report folder
def storeIncrementValue():
    global autoIncrement
    global dirpath
    file_path = dirpath.joinpath('increment.txt')
    if file_path.is_file():
        with open (file_path,'w') as f:  
            f.write(str(autoIncrement))
            #print('File was updated for auto increment')
    else:
        with open (file_path,'w') as f:  
            f.write(str(autoIncrement))
            #print('File was created for auto increment')

#To store the done list detail for each group in local report folder
def storeDoneList(doneListName,doneList):
    global dirpath
    file_path = dirpath.joinpath(doneListName+'.txt')
    if file_path.is_file():
        with open (file_path,'w') as f:  
            f.write(str(doneList))
            #print('File was updated for '+doneListName)
    else:
        with open (file_path,'w') as f:  
            f.write(str(doneList))
            #print('File was created for '+doneListName)

#To retrieve the done list of the day for each group from local report folder          
def getDoneListValue():
    global equityDone
    global futureDone
    global optionDone
    global indexFutureDone
    global indexOptionDone
    global commodityDone
    global dirpath
    file_path = dirpath.joinpath('equityDone.txt')
    todayDate=datetime.datetime.now()
    if file_path.is_file():
        fileTime=datetime.datetime.fromtimestamp(os.path.getmtime(file_path))
        if(todayDate.day==fileTime.day and todayDate.month==fileTime.month and todayDate.year==fileTime.year):
            with open (file_path,'r') as f:
                equityDone= json.loads(str(f.read()).replace("'",'"'))
                #print("Done list was loaded from text file for equityDone")
           
    file_path = dirpath.joinpath('futureDone.txt')
    if file_path.is_file():
        fileTime=datetime.datetime.fromtimestamp(os.path.getmtime(file_path))
        if(todayDate.day==fileTime.day and todayDate.month==fileTime.month and todayDate.year==fileTime.year):
            with open (file_path,'r') as f:
                futureDone= json.loads(str(f.read()).replace("'",'"'))
                #print("Done list was loaded from text file for equityFurtueDone")
     
    file_path = dirpath.joinpath('optionDone.txt')
    if file_path.is_file():
        fileTime=datetime.datetime.fromtimestamp(os.path.getmtime(file_path))
        if(todayDate.day==fileTime.day and todayDate.month==fileTime.month and todayDate.year==fileTime.year):
            with open (file_path,'r') as f:
                optionDone= json.loads(str(f.read()).replace("'",'"'))
                #print("Done list was loaded from text file for optionDone")
   
    file_path = dirpath.joinpath('indexFutureDone.txt')
    if file_path.is_file():
        fileTime=datetime.datetime.fromtimestamp(os.path.getmtime(file_path))
        if(todayDate.day==fileTime.day and todayDate.month==fileTime.month and todayDate.year==fileTime.year):
            with open (file_path,'r') as f:
                indexFutureDone= json.loads(str(f.read()).replace("'",'"'))
                #print("Done list was loaded from text file for indexFutureDone")
           
    file_path = dirpath.joinpath('indexOptionDone.txt')
    if file_path.is_file():
        fileTime=datetime.datetime.fromtimestamp(os.path.getmtime(file_path))
        if(todayDate.day==fileTime.day and todayDate.month==fileTime.month and todayDate.year==fileTime.year):
            with open (file_path,'r') as f:
                indexOptionDone= json.loads(str(f.read()).replace("'",'"'))
                #print("Done list was loaded from text file for indexOptionDone")
           
    file_path = dirpath.joinpath('commodityDone.txt')
    if file_path.is_file():
        fileTime=datetime.datetime.fromtimestamp(os.path.getmtime(file_path))
        if(todayDate.day==fileTime.day and todayDate.month==fileTime.month and todayDate.year==fileTime.year):
            with open (file_path,'r') as f:
                commodityDone= json.loads(str(f.read()).replace("'",'"'))
                #print("Done list was loaded from text file for commodityDone")
           
getDoneListValue()

#To store the execute order details in local report folder as a summary for each day          
def storeOutput(resultList):
    try:
        global dirpath
        date1=datetime.datetime.now()
        name=str(date1.day)+'_'+str(date1.month)+'_'+str(date1.year)+'_report.xlsx'
        file_path = dirpath.joinpath(name)
        if(len(resultList)>0):
            if file_path.is_file():
                wb=openpyxl.load_workbook(file_path)
                ws=wb.worksheets[0]
                row=ws.max_row
                for s in resultList:
                    row=row+1
                    queryParam=s.url.split("?")[1].split("&")
                    for param in queryParam:
                        keyValue=param.split('=')
                        if(keyValue[0]=='UniqueID'):
                            ws.cell(row,1).value=keyValue[1]
                        elif(keyValue[0]=='SourceSymbol'):
                            ws.cell(row,2).value=keyValue[1]
                        elif(keyValue[0]=='StrategyTag'):
                            ws.cell(row,3).value=keyValue[1]
                        elif(keyValue[0]=='TransactionType'):
                            ws.cell(row,4).value=keyValue[1]
                        elif(keyValue[0]=='OptionType'):
                            ws.cell(row,5).value=keyValue[1]
                        elif(keyValue[0]=='Quantity'):
                            ws.cell(row,6).value=keyValue[1]
                    ws.cell(row,7).value=s.content
                    ws.cell(row,8).value=date1
                wb.save(file_path)
                print('Report was updated for executed orders')
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.cell(1,1).value='UniqueID'
                ws.cell(1,2).value='Symbol'
                ws.cell(1,3).value='StrategyTag'
                ws.cell(1,4).value='TransactionType'
                ws.cell(1,5).value='OptionType'
                ws.cell(1,6).value='Quantity'
                ws.cell(1,7).value='Response'
                ws.cell(1,8).value='Date and Time'
                row=1
                for s in resultList:
                    row=row+1
                    queryParam=s.url.split("?")[1].split("&")
                    for param in queryParam:
                        keyValue=param.split('=')
                        if(keyValue[0]=='UniqueID'):
                            ws.cell(row,1).value=keyValue[1]
                        elif(keyValue[0]=='SourceSymbol'):
                            ws.cell(row,2).value=keyValue[1]
                        elif(keyValue[0]=='StrategyTag'):
                            ws.cell(row,3).value=keyValue[1]
                        elif(keyValue[0]=='TransactionType'):
                            ws.cell(row,4).value=keyValue[1]
                        elif(keyValue[0]=='OptionType'):
                            ws.cell(row,5).value=keyValue[1]
                        elif(keyValue[0]=='Quantity'):
                            ws.cell(row,6).value=keyValue[1]
                    ws.cell(row,7).value=s.content
                    ws.cell(row,8).value=date1
                wb.save(file_path)
                print('Report was created for executed orders')
    except:
        print("Error while storing the summary in report")
       

#To retrieve the unique id from local report folder
def getIncrementValue():
    global autoIncrement
    global dirpath
    file_path = dirpath.joinpath('increment.txt')
    if file_path.is_file():
        with open (file_path,'r') as f:
            autoIncrement= int(f.read())
    else:
        autoIncrement=0
    #print('AutoIncrement value readed from txt file: '+str(autoIncrement))
   

def exception_handler1(request,exception):
    print("exception occured", request, exception)

#To execute the list of orders based on group
def executeOrder(groupdic,datas,appendLast):
    global async_list
    for data in datas:
        symbols=data["Name"]+appendLast
        if(symbols is None):
            break
        global userName
        global autoIncrement
        url=None
        if(data["BUY"]=="Yes" ):
            for i in groupdic:
                autoIncrement=autoIncrement+1
                url="http://localhost:21000/MappedOrderAdv?StrategyTag="+i+"&UniqueID="+str(autoIncrement)+"&SourceSymbol="+symbols+"&TransactionType=LE&OrderType=MARKET&Quantity="+str(groupdic[i])
                if(url is not None):
                    async_list.append(grequests.get(url))
        elif(data["EXIT BUY"]=="Yes"):
            for i in groupdic:
                autoIncrement=autoIncrement+1
                url="http://localhost:21000/MappedOrderAdv?StrategyTag="+i+"&UniqueID="+str(autoIncrement)+"&SourceSymbol="+symbols+"&TransactionType=LX&OrderType=MARKET&Quantity="+str(groupdic[i])
                if(url is not None):
                    async_list.append(grequests.get(url))
        elif(data["SELL"]=="Yes"):
            for i in groupdic:
                autoIncrement=autoIncrement+1
                url="http://localhost:21000/MappedOrderAdv?StrategyTag="+i+"&UniqueID="+str(autoIncrement)+"&SourceSymbol="+symbols+"&TransactionType=SE&OrderType=MARKET&Quantity="+str(groupdic[i])
                if(url is not None):
                    async_list.append(grequests.get(url))
        elif(data["EXIT SELL"]=="Yes"):
            for i in groupdic:
                autoIncrement=autoIncrement+1
                url="http://localhost:21000/MappedOrderAdv?StrategyTag="+i+"&UniqueID="+str(autoIncrement)+"&SourceSymbol="+symbols+"&TransactionType=SX&OrderType=MARKET&Quantity="+str(groupdic[i])
                if(url is not None):
                    async_list.append(grequests.get(url))

#To execute the list of option orders based on group
def executeOrderOpt(groupdic,datas,appendLast):
    global async_list
    for data in datas:
        name=data["Name"].split("_")
        symbols=name[0]+appendLast
        if(symbols is None):
            break
        global userName
        global autoIncrement
        url=None
        if(data["BUY"]=="Yes" ):
            for i in groupdic:
                autoIncrement=autoIncrement+1
                url="http://localhost:21000/MappedOrderAdv?StrategyTag="+i+"&UniqueID="+str(autoIncrement)+"&SourceSymbol="+symbols+"&TransactionType=LE&OptionType="+data["TransactionType"]+"&OrderType=MARKET&Quantity="+str(groupdic[i])
                if(url is not None):
                    async_list.append(grequests.get(url))
        elif(data["EXIT BUY"]=="Yes"):
            for i in groupdic:
                autoIncrement=autoIncrement+1
                url="http://localhost:21000/MappedOrderAdv?StrategyTag="+i+"&UniqueID="+str(autoIncrement)+"&SourceSymbol="+symbols+"&TransactionType=LX&OptionType="+data["TransactionType"]+"&OrderType=MARKET&Quantity="+str(groupdic[i])
                if(url is not None):
                    async_list.append(grequests.get(url))
        elif(data["SELL"]=="Yes"):
            for i in groupdic:
                autoIncrement=autoIncrement+1
                url="http://localhost:21000/MappedOrderAdv?StrategyTag="+i+"&UniqueID="+str(autoIncrement)+"&SourceSymbol="+symbols+"&TransactionType=LE&OptionType="+data["TransactionType"]+"&OrderType=MARKET&Quantity="+str(groupdic[i])
                if(url is not None):
                    async_list.append(grequests.get(url))
        elif(data["EXIT SELL"]=="Yes"):
            for i in groupdic:
                autoIncrement=autoIncrement+1
                url="http://localhost:21000/MappedOrderAdv?StrategyTag="+i+"&UniqueID="+str(autoIncrement)+"&SourceSymbol="+symbols+"&TransactionType=LX&OptionType="+data["TransactionType"]+"&OrderType=MARKET&Quantity="+str(groupdic[i])
                if(url is not None):
                    async_list.append(grequests.get(url))

#To check whether csv was updated to read
def getExcelUpdated(name, updatedTime):
    try:
        global excelurl
        excelpath = Path(excelurl)
        file_path = excelpath.joinpath(name+'.csv')
        m=os.path.getmtime(file_path)
        if(m==updatedTime):
            #print(name+" excel was not updated")
            return False,updatedTime
        else:
            print(name+" excel was updated")
            return True,m
    except:
        return False,updatedTime

#To read the symbols and transaction type from each csv files
def readExcel(name, doneList, startFrom):
    #Reading the excel
    try:
        global excelurl
        excelpath = Path(excelurl)
        file_path = excelpath.joinpath(name+'.csv')
       
        dff = pd.read_csv(file_path)
        df = pd.DataFrame(dff)
       
        datas=[]
        inc=0
        if(startFrom>=len(df)):
            return datas,startFrom
        for i in range(startFrom,len(df)):
            c=df.loc[i].copy()
            nameSplit = c["Name"].strip().split(" ")
            name=nameSplit[0]
            c["Name"]=name
            if(len(nameSplit)==1):
                if(name is not None and name != ''):
                    if(c["BUY"]=="Yes"):
                        if((name not in doneList["BUY"] or((name in doneList["EXIT BUY"]) and doneList["BUY"][name]==doneList["EXIT BUY"][name])) and (name not in doneList["SELL"] or(name in doneList["EXIT SELL"] and doneList["SELL"][name]==doneList["EXIT SELL"][name]))):
                            datas.append(c)
                            doneList=updateValue(doneList, c)
                        elif((name in doneList["SELL"] and(name not in doneList["EXIT SELL"] or doneList["SELL"][name]>doneList["EXIT SELL"][name]))):
                            k=c.copy()
                            k["BUY"]=None
                            k["EXIT SELL"]="Yes"
                            datas.append(k)
                            doneList=updateValue(doneList, k)
                    elif(c["SELL"]=="Yes"):
                        if((name not in doneList["SELL"] or((name in doneList["EXIT SELL"]) and doneList["SELL"][name]==doneList["EXIT SELL"][name])) and (name not in doneList["BUY"] or(name in doneList["EXIT BUY"] and doneList["BUY"][name]==doneList["EXIT BUY"][name]))):
                            datas.append(c)
                            doneList=updateValue(doneList, c)
                        elif((name in doneList["BUY"] and(name not in doneList["EXIT BUY"] or doneList["BUY"][name]>doneList["EXIT BUY"][name]))):
                            k=c.copy()
                            k["SELL"]=None
                            k["EXIT BUY"]="Yes"
                            datas.append(k)
                            doneList=updateValue(doneList, k)
                    elif(c["EXIT BUY"]=="Yes"):
                        if((name in doneList["BUY"]) and((name not in doneList["EXIT BUY"]) or doneList["BUY"][name]>doneList["EXIT BUY"][name])):
                            datas.append(c)
                            doneList=updateValue(doneList, c)
                        elif((name in doneList["SELL"] and(name not in doneList["EXIT SELL"] or doneList["SELL"][name]>doneList["EXIT SELL"][name]))):
                            k=c.copy()
                            k["EXIT BUY"]=None
                            k["EXIT SELL"]="Yes"
                            datas.append(k)
                            doneList=updateValue(doneList, k)
                    elif(c["EXIT SELL"]=="Yes"):
                        #print(name,doneList)
                        if((name in doneList["SELL"]) and((name not in doneList["EXIT SELL"]) or doneList["SELL"][name]>doneList["EXIT SELL"][name])):
                            datas.append(c)
                            doneList=updateValue(doneList, c)
                        elif((name in doneList["BUY"] and(name not in doneList["EXIT BUY"] or doneList["BUY"][name]>doneList["EXIT BUY"][name]))):
                            k=c.copy()
                            k["EXIT SELL"]=None
                            k["EXIT BUY"]="Yes"
                            datas.append(k)
                            doneList=updateValue(doneList, k)
                inc=inc+1
        return datas,startFrom+inc,doneList
    except:
        return [],startFrom,doneList
    
    #Reading the excel for Options
def readExcelOpt(name, doneList, startFrom):
    try:
        global excelurl
        excelpath = Path(excelurl)
        file_path = excelpath.joinpath(name+'.csv')
       
        dff = pd.read_csv(file_path)
        df = pd.DataFrame(dff)
       
        datas=[]
        inc=0
        if(startFrom>=len(df)):
            return datas,startFrom
        for i in range(startFrom,len(df)):
            c=df.loc[i].copy()
            nameSplit = c["Name"].strip().split(" ")          
            if(len(nameSplit)>1):
                name=nameSplit[0]+"_"+nameSplit[4]
                c["Name"]=name
                c["TransactionType"]=nameSplit[4]
                if(name is not None and name != ''):
                    if(c["BUY"]=="Yes"):
                        if((name not in doneList["BUY"] or((name in doneList["EXIT BUY"]) and doneList["BUY"][name]==doneList["EXIT BUY"][name])) and (name not in doneList["SELL"] or(name in doneList["EXIT SELL"] and doneList["SELL"][name]==doneList["EXIT SELL"][name]))):
                            datas.append(c)
                            doneList=updateValue(doneList, c)
                    elif(c["SELL"]=="Yes"):
                        if((name in doneList["BUY"] and(name not in doneList["EXIT BUY"] or doneList["BUY"][name]>doneList["EXIT BUY"][name]))):
                            k=c.copy()
                            k["SELL"]=None
                            k["EXIT BUY"]="Yes"
                            datas.append(k)
                            doneList=updateValue(doneList, k)
                    elif(c["EXIT BUY"]=="Yes"):
                        if((name in doneList["BUY"]) and((name not in doneList["EXIT BUY"]) or doneList["BUY"][name]>doneList["EXIT BUY"][name])):
                            datas.append(c)
                            doneList=updateValue(doneList, c)
                    elif(c["EXIT SELL"]=="Yes"):
                        if((name in doneList["BUY"] and(name not in doneList["EXIT BUY"] or doneList["BUY"][name]>doneList["EXIT BUY"][name]))):
                            k=c.copy()
                            k["EXIT SELL"]=None
                            k["EXIT BUY"]="Yes"
                            datas.append(k)
                            doneList=updateValue(doneList, k)
                inc=inc+1
        return datas,startFrom+inc,doneList
    except:
        return [],startFrom,doneList

#To update the done value for validation purpose
def updateValue(doneList,c):
    #print(doneList,data)
    if  c is not None :
        name=c["Name"]
        #print(name)
        if(c["BUY"]=="Yes"):
            if(name in doneList["BUY"]):
                doneList["BUY"][name]=doneList["BUY"][name]+1
            else:
                doneList["BUY"][name]=1
        elif(c["SELL"]=="Yes"):
            if(name in doneList["SELL"]):
                doneList["SELL"][name]=doneList["SELL"][name]+1
            else:
                doneList["SELL"][name]=1
        elif(c["EXIT BUY"]=="Yes"):
            if(name in doneList["EXIT BUY"]):
                doneList["EXIT BUY"][name]=doneList["EXIT BUY"][name]+1
            else:
                doneList["EXIT BUY"][name]=1
        elif(c["EXIT SELL"]=="Yes"):
            if(name in doneList["EXIT SELL"]):
                doneList["EXIT SELL"][name]=doneList["EXIT SELL"][name]+1
            else:
                doneList["EXIT SELL"][name]=1
    return doneList

#To retrieve group detail from master excel sheet
def getGroup():
    global excelurl
    #Reading the excel
    excelpath = Path(excelurl)
    file_path = excelpath.joinpath('Stoxxo-Master Sheet.xlsx')
    wb=openpyxl.load_workbook(file_path)
    #print(wb)
    sheet=wb["Sheet1"]
    maxcolumn=sheet.max_column
    maxrow=sheet.max_row
    #print(maxcolumn)
    global equityGroups
    global futureGroups
    global optionGroups
    global indexFutureGroups
    global indexOptionGroups
    global commodityGroups
    equityGroups={}
    futureGroups={}
    optionGroups={}
    indexFutureGroups={}
    indexOptionGroups={}
    commodityGroups={}
    colum={}
    for i in range(3,maxcolumn+1):
        if(sheet.cell(2,i).value!=None):
            colum[i]=sheet.cell(2,i).value
        else:
            maxcolumn=i-1
            break
    #print(colum)
    colum2={}
   
    for i in range(3,maxcolumn+1):
        colum2[sheet.cell(2,i).value]=i
    #print(colum2)
    for i in colum:
        co=colum[i]
        #print(co,sheet.cell(3,i).value,sheet.cell(10,colum2[co]).value)
        if(sheet.cell(3,i).value==1 and sheet.cell(11,colum2[co]) !=None and sheet.cell(11,colum2[co]).value>0):
            equityGroups[co]=sheet.cell(11,colum2[co]).value
        if(sheet.cell(4,i).value==1 and sheet.cell(12,colum2[co]) !=None and sheet.cell(12,colum2[co]).value>0):
            futureGroups[co]=sheet.cell(12,colum2[co]).value
        if(sheet.cell(5,i).value==1 and sheet.cell(13,colum2[co]) !=None and sheet.cell(13,colum2[co]).value>0):
            optionGroups[co]=sheet.cell(13,colum2[co]).value
        if(sheet.cell(6,i).value==1 and sheet.cell(14,colum2[co]) !=None and sheet.cell(14,colum2[co]).value>0):
            indexFutureGroups[co]=sheet.cell(14,colum2[co]).value
        if(sheet.cell(7,i).value==1 and sheet.cell(15,colum2[co]) !=None and sheet.cell(15,colum2[co]).value>0):
            indexOptionGroups[co]=sheet.cell(15,colum2[co]).value
        if(sheet.cell(8,i).value==1 and sheet.cell(16,colum2[co]) !=None and sheet.cell(16,colum2[co]).value>0):
            commodityGroups[co]=sheet.cell(16,colum2[co]).value
   
    global MCXTiming
    global NSETiming
    try:
        if(sheet.cell(20,3)!=None and sheet.cell(21,3)!=None):
            a=str(sheet.cell(20,3).value).split(":")
            NSETiming["start"]["hour"]=int(a[0])
            NSETiming["start"]["min"]=int(a[1])
            a=str(sheet.cell(21,3).value).split(":")
            NSETiming["end"]["hour"]=int(a[0])
            NSETiming["end"]["min"]=int(a[1])

        if(sheet.cell(23,3)!=None and sheet.cell(24,3)!=None):
            a=str(sheet.cell(23,3).value).split(":")
            MCXTiming["start"]["hour"]=int(a[0])
            MCXTiming["start"]["min"]=int(a[1])
            a=str(sheet.cell(24,3).value).split(":")
            MCXTiming["end"]["hour"]=int(a[0])
            MCXTiming["end"]["min"]=int(a[1])
    except:
        MCXTiming={"start":{"hour":9,"min":16},"end":{"hour":15,"min":25}}
        NSETiming={"start":{"hour":9,"min":1},"end":{"hour":23,"min":25}}
       

#Method to read excel and execute orders in different thread at regular
def diffThread():
    getGroup()
    i=1
    nseExitPlaced=0
    mcxExitPlaced=0
    while(i):
        global exitDay
        date=datetime.datetime.now()
        if(date>=exitDay):
            #print("Trading subscription got expired")
            global runningGui
            runningGui.destroy()
            sys.exit()
        getIncrementValue()
        global refreshRate
        global equityDone
        global equityDoneRow
        global futureDone
        global futureDoneRow
        global optionDone
        global optionDoneRow
        global indexFutureDone
        global indexFutureDoneRow
        global indexOptionDone
        global indexOptionDoneRow
        global commodityDone
        global commodityDoneRow
        global equityUpdated
        global futureUpdated
        global optionUpdated
        global indexFutureUpdated
        global indexOptionUpdated
        global commodityUpdated
        global MCXTiming
        global NSETiming

       
        global async_list
        #Initializing variables at the start of the month
        if(date>date.replace(hour=6,minute=0,second=0) and date<date.replace(hour=6,minute=0,second=30)):
            nseExitPlaced=0
            mcxExitPlaced=0
            getGroup()
            equityDone={"BUY":{},"SELL":{},"EXIT BUY":{},"EXIT SELL":{}}
            futureDone={"BUY":{},"SELL":{},"EXIT BUY":{},"EXIT SELL":{}}
            optionDone={"BUY":{},"SELL":{},"EXIT BUY":{},"EXIT SELL":{}}
            indexFutureDone={"BUY":{},"SELL":{},"EXIT BUY":{},"EXIT SELL":{}}
            indexOptionDone={"BUY":{},"SELL":{},"EXIT BUY":{},"EXIT SELL":{}}
            commodityDone={"BUY":{},"SELL":{},"EXIT BUY":{},"EXIT SELL":{}}

        equityDoneRow=0
        futureDoneRow=0
        optionDoneRow=0
        indexFutureDoneRow=0
        indexOptionDoneRow=0
        commodityDoneRow=0
        async_list=[]
        global startProcess
        if (startProcess!=True):
            break
        #Execute orders after every 2 seconds
        if(date>date.replace(hour=NSETiming["start"]["hour"],minute=NSETiming["start"]["min"],second=0) and date<date.replace(hour=NSETiming["end"]["hour"],minute=NSETiming["end"]["min"],second=0)):
            m,equityUpdated=getExcelUpdated("Equity",equityUpdated)
            if(m==True and len(equityGroups)>0):
                data,equityDoneRow,equityDone=readExcel("Equity",equityDone,equityDoneRow)
                executeOrder(equityGroups,data,'_SIM')
                storeDoneList("equityDone",equityDone)
        if(date>date.replace(hour=NSETiming["start"]["hour"],minute=NSETiming["start"]["min"],second=0) and date<date.replace(hour=NSETiming["end"]["hour"],minute=NSETiming["end"]["min"],second=0)):
            m,futureUpdated=getExcelUpdated("Equity",futureUpdated)
            if(m==True and len(futureGroups)>0):
                data,futureDoneRow,futureDone=readExcel("Equity",futureDone,futureDoneRow)
                executeOrder(futureGroups,data,'_FUT')
                storeDoneList("futureDone",futureDone)
        if(date>date.replace(hour=NSETiming["start"]["hour"],minute=NSETiming["start"]["min"],second=0) and date<date.replace(hour=NSETiming["end"]["hour"],minute=NSETiming["end"]["min"],second=0)):
            m,optionUpdated=getExcelUpdated("Equity",optionUpdated)
            if(m==True and len(optionGroups)>0):
                data,optionDoneRow,optionDone=readExcelOpt("Equity",optionDone,optionDoneRow)
                executeOrderOpt(optionGroups,data,'_OPT')
                storeDoneList("optionDone",optionDone)
        if(date>date.replace(hour=NSETiming["start"]["hour"],minute=NSETiming["start"]["min"],second=0) and date<date.replace(hour=NSETiming["end"]["hour"],minute=NSETiming["end"]["min"],second=0)):
            m,indexFutureUpdated=getExcelUpdated("Index",indexFutureUpdated)
            if(m==True and len(indexFutureGroups)>0):
                data,indexFutureDoneRow,indexFutureDone=readExcel("Index",indexFutureDone,indexFutureDoneRow)
                executeOrder(indexFutureGroups,data,'_FUT')
                storeDoneList("indexFutureDone",indexFutureDone)
        if(date>date.replace(hour=NSETiming["start"]["hour"],minute=NSETiming["start"]["min"],second=0) and date<date.replace(hour=NSETiming["end"]["hour"],minute=NSETiming["end"]["min"],second=0)):
            m,indexOptionUpdated=getExcelUpdated("Index",indexOptionUpdated)
            if(m==True and len(indexOptionGroups)>0):
                data,indexOptionDoneRow,indexOptionDone=readExcelOpt("Index",indexOptionDone,indexOptionDoneRow)
                executeOrderOpt(indexOptionGroups,data,'_OPT')
                storeDoneList("indexOptionDone",indexOptionDone)
        if(date>date.replace(hour=MCXTiming["start"]["hour"],minute=MCXTiming["start"]["min"],second=0) and date<date.replace(hour=MCXTiming["end"]["hour"],minute=MCXTiming["end"]["min"],second=0)):
            m,commodityUpdated=getExcelUpdated("Commodity",commodityUpdated)
            if(m==True and len(commodityGroups)>0):
                data,commodityDoneRow,commodityDone=readExcel("Commodity",commodityDone,commodityDoneRow)
                executeOrder(commodityGroups,data,'_FUT')
                storeDoneList("commodityDone",commodityDone)
           
        output=grequests.map(async_list,exception_handler=exception_handler1)
        print('No. of orders placed at this iteration: ' +str(len(output)))
        storeOutput(output)
        storeIncrementValue()
        sleep(refreshRate)
   
def start():
    t=threading.Thread(target=diffThread)
    t.start()
   
#Gui for starting and stoping trading
def gui2():
    global runningGui
    global startProcess
    global stopProcess
    root=Tk()
    runningGui=root
    def click():
        global startProcess
        global stopProcess
        startProcess=True
        stopProcess=False
        root.destroy()
        start()
        gui2()
       
    def cancel():
        root.destroy()
        global startProcess
        global stopProcess
        startProcess=False
        stopProcess=True

    def exit1():
        root.destroy()
        global startProcess
        global stopProcess
        startProcess=False
        stopProcess=True
    def stop():
        root.destroy()
        global startProcess
        global stopProcess
        startProcess=False
        stopProcess=True
        gui2()
    def roundOff():
        answer=ms.askyesno(title='Confirmation',message='Are you sure that you want to perform square off for all the users?')
        if answer:
            import requests
            root.destroy()
            global stopProcess
            stopProcess=False      
            result=requests.get("http://localhost:21000/SquareOffAll")
            print(result.content)
            gui2()
    def roundOff2():
        ms.showerror(title='Square off Error', message='Please stop the trading to perform square off')
       

    root.title(' TIGER NIFTY ')
    Label(root,text=' ',font='none 10').grid(row=1,column=2,sticky=W)
    Label(root,text=' ',font='none 20').grid(row=2,column=3,sticky=W)
    Label(root,text=' ',font='none 10').grid(row=4,column=0,sticky=W)
    Label(root,text=' ',font='none 10').grid(row=4,column=2,sticky=W)
    Label(root,text=' ',font='none 10').grid(row=4,column=4,sticky=W)
    Label(root,text='          ',font='none 10').grid(row=4,column=0,sticky=W)
    Label(root,text='          ',font='none 10').grid(row=4,column=3,sticky=W)
    Label(root,text='          ',font='none 10').grid(row=4,column=7,sticky=W)
   
    Label(root,text='       ',font='none 30').grid(row=4,column=4,sticky=W)
    if(startProcess==True):
       
        Label(root,text='Running process...',font='none 10').grid(row=4,column=1,sticky=E)
        Button(root,text='Stop',width=10,command=stop).grid(row=4,column=5,sticky=E)
        Button(root,text='Square off',width=10,command=roundOff2).grid(row=6,column=5,sticky=E)
        Button(root,text='Exit',width=10,command=exit1).grid(row=6,column=1,sticky=E)
    else:
        Button(root,text='Start',width=10,command=click).grid(row=4,column=1,sticky=E)
        Label(root,text=' ',font='none 10').grid(row=4,column=4,sticky=E)
        Button(root,text='Exit',width=10,command=exit1).grid(row=6,column=3,sticky=E)
    if(stopProcess==True):
        Button(root,text='Square off',width=10,command=roundOff).grid(row=4,column=5,sticky=E)
    Label(root,text=' ',font='none 10').grid(row=5,column=0,sticky=W)
    Label(root,text='          ',font='none 10').grid(row=6,column=2,sticky=W)
    Label(root,text='          ',font='none 10').grid(row=7,column=2,sticky=W)
    Label(root,text='          ',font='none 10').grid(row=8,column=2,sticky=W)
    root.mainloop()

def userLogin(userId):
    gui2()
   
#Gui to select the folder for reading csv and proceed
def gui():
    root=Tk()
    root.title(' TIGER NIFTY ')
    Label(root,text='        Welcome to tiger nifty',font='none 10').grid(row=1,column=1,sticky=W)
    global excelurl
    def click():
        excelpath = Path(excelurl)
        file_path = excelpath.joinpath('Stoxxo-Master Sheet.xlsx')
        if file_path.is_file():
            try:
                import requests  
                result=requests.get("http://localhost:21000")
                root.destroy()
                gui2()
            except:
                ms.showerror(title='Stoxxo Error', message='Please start trading in stoxxo to continue')
        else:
            ms.showerror(title='Master file not found', message='Please provide valid directory to continue')
    def browse():
        global excelurl
        root.directory=filedialog.askdirectory()
        excelurl=root.directory  
        root.destroy()
        gui()
           
    def cancel():
        root.destroy()
    Label(root,text=' ',font='none 10').grid(row=2,column=0,sticky=W)
    Label(root,text=' ',font='none 10').grid(row=3,column=0,sticky=W)
    Label(root,text='Choosen file path:  ',font='none 10').grid(row=4,column=0,sticky=W)
    Label(root,text=excelurl,font='none 10').grid(row=4,column=1,sticky=W)
    Label(root,text=' ',font='none 10').grid(row=5,column=0,sticky=W)
    Button(root,text='Browse',width=6,command=browse).grid(row=6,column=0)
    Label(root,text='          ',font='none 10').grid(row=6,column=1,sticky=W)
    Button(root,text='Continue',width=6,command=click).grid(row=6,column=2,sticky=E)
    Label(root,text='          ',font='none 10').grid(row=6,column=3,sticky=W)
    Label(root,text='          ',font='none 10').grid(row=7,column=3,sticky=W)
    root.mainloop()
gui()

